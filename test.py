from progress.bar import Bar
import os
from PIL import Image
from io import BytesIO
import requests
from openpyxl import load_workbook
from fpdf import FPDF
import tempfile


def download_images(image_links):
    try:
        os.system("cls")
        images = []
        for link in Bar("Downloading").iter(image_links):
            try:
                response = requests.get(link)
                if response.status_code == 200:
                    images.append(response.content)
                else:
                    print(f"Failed to download: {link}")
            except Exception as e:
                print(f"Error occurred while downloading {link}: {str(e)}")
        return images
    except Exception:
        print(f"Error for link - {link}")


def create_pdf(image_data):
    pdf = FPDF()
    temp_path_images = []
    for data in image_data:
        image_stream = BytesIO(data)
        image = Image.open(image_stream)
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".jpg")
        temp_file_path = temp_file.name
        image.save(temp_file_path, format="JPEG")
        temp_path_images.append(temp_file_path)
        pdf.add_page()
        pdf.image(temp_file_path, 0, 0)
    return [temp_path_images, pdf]


# Example usage:
excel_file = "image_links.xlsx"

# Load the Excel file and read the image links
workbook = load_workbook(excel_file)
sheet = workbook.active
image_links = [cell.value for row in sheet.iter_rows() for cell in row]
# print(image_links)

# Download the images
image_data = download_images(image_links)
# print(len())

# Create the PDF
temp_file_paths, pdf = create_pdf(image_data)

# Save the PDF file
pdf_file = "combined_images.pdf"
pdf.output(pdf_file)
print(f"PDF file '{pdf_file}' created successfully.")
for path in temp_file_paths:
    os.remove(path)
